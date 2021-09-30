import pymysql
import os
import zipfile
import csv
import sys
import logging
import time
import datetime
import shutil
import numpy as np
from openpyxl.reader.excel import load_workbook
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import xlsxwriter

start_date_str =input("Please enter start date, ex:2018-3-1:\n")
end_date_str = input("Please enter end date, ex:2019-9-30:\n")
start_date_str = str(start_date_str) + " 0:0:0"
end_date_str = str(end_date_str) + " 23:55:0"
start_date = datetime.strptime(start_date_str, "%Y-%m-%d %H:%M:%S")
end_date = datetime.strptime(end_date_str, "%Y-%m-%d %H:%M:%S")
total_serial = int(((end_date-start_date).total_seconds()/60)/5) + 1
str1="2018-3-1 00:00:00"
start = datetime.strptime(str1, "%Y-%m-%d %H:%M:%S")
total_list = [ [None] * total_serial for i in range(8) ]

path = "D:/te/B船_KYMA"
for root,dirs,files in os.walk(path):
    for name in files:
        if name.startswith('YMUpsurgence'):
            tmp = name.split("_")
            type = tmp[len(tmp)-2].strip()
            if type == "torque" or type == "thrust" or type == "engine speed" or type == "power" or type == "energy" or type == "total revolutions":
                day = tmp[len(tmp)-1].strip(".txt")
                file_path = os.path.join(root, name)
                filecontent = open(file_path,'r')
                for lines in filecontent.readlines():
                    lines = lines.strip()
                    lines_split = lines.split(",")
                    record_time = datetime.strptime(lines_split[0], "%Y-%m-%d %H:%M:%S")
                    if record_time <= end_date:
                        serial = int(((record_time-start_date).total_seconds()/60)/5)
                        num=int(((record_time-start).total_seconds()/60)/5)

                        if type == "torque":
                            total_list[7][serial] = str(lines_split[0])
                            total_list[1][serial] = str(lines_split[1])
                            total_list[0][serial] = str(num)
                        elif type == "thrust":
                            total_list[7][serial] = str(lines_split[0])
                            total_list[2][serial] = str(lines_split[1])
                        elif type == "engine speed":
                            total_list[7][serial] = str(lines_split[0])
                            total_list[3][serial] = str(lines_split[1])
                        elif type == "power":
                            total_list[7][serial] = str(lines_split[0])
                            total_list[4][serial] = str(lines_split[1])
                        elif type == "energy":
                            total_list[7][serial] = str(lines_split[0])
                            total_list[5][serial] = str(lines_split[1])
                        elif type == "total revolutions":
                            total_list[7][serial] = str(lines_split[0])
                            total_list[6][serial] = str(lines_split[1])
                        else:
                            pass
print("len(total_list): {0}, len(total_list[0]): {1}".format(len(total_list), len(total_list[0])))
inverse_total_list =list( map(list, zip(*total_list)))

total_name = "B.csv"
file_path = os.path.join(path, total_name)
workbook = xlsxwriter.Workbook(file_path)
worksheet = workbook.add_worksheet()
row = 0
col = 0
worksheet.write(0, 0, "number")
worksheet.write(0, 1, "torque")
worksheet.write(0, 2, "thrust")
worksheet.write(0, 3, "engine speed")
worksheet.write(0, 4, "power")
worksheet.write(0, 5, "energy")
worksheet.write(0, 6, "total revolutions")
worksheet.write(0, 7, "datatime")
for i in range(len(inverse_total_list)):
    for j in range(len(inverse_total_list[i])):
        worksheet.write(i+1, j, str(inverse_total_list[i][j]))
workbook.close()
